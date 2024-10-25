import os
import openpyxl.workbook
import pdfplumber
import re
from datetime import datetime

#vendo a quantidade de imagens e nome da planilha

directory = 'pdf'
files = os.listdir(directory)
files_quantity = len(files)

if files_quantity == 0:
    raise Exception("Sem arquivos na pasta")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'importacoes'

#dados que nós queremos registrar

ws['A1'] = 'importacoes #'
ws['B1'] = 'Valor'
ws['C1'] = 'File name'
ws['D1'] = 'Status'

last_empty_line = 1

while ws['A' + str(last_empty_line)].value is not None:
    last_empty_line += 1

for file in files:
    with pdfplumber.open(directory + "/" + file) as pdf:
        first_page = pdf.pages[0]
        pdf_text = first_page.extract_text()
        print(pdf_text)

    arquiv_agencia_re_pattern = r'Agência: (\d+)'
    arquiv_valor_re_pattern = r'Valor: R(\d+)'

    match_agencia = re.search(arquiv_agencia_re_pattern, pdf_text)
    match_valor = re.search(arquiv_valor_re_pattern, pdf_text)

    if match_agencia:
        import_agencia = match_agencia.group(1)
        ws['A{}'.format(last_empty_line)] = import_agencia
    else:
        ws['A{}'.format(last_empty_line)] = "não consegui achar o número da agencia"

    if match_valor:
        import_valor = match_valor.group(1)
        ws['B{}'.format(last_empty_line)] = import_valor
    else:
        ws['B{}'.format(last_empty_line)] = "não consegui achar o valor"

    ws['C{}'.format(last_empty_line)] = file
    ws['D{}'.format(last_empty_line)] = "Completed"

    last_empty_line += 1

full_now = str(datetime.now()).replace(":", "-")
dot_index = full_now.index(".")
now = full_now[:dot_index]

wb.save("Imports - {}.xlsx".format(now))


